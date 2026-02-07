import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import tempfile

st.set_page_config(
    page_title="CBSE Ratio Calculator",
    page_icon="📊",
    layout="wide"
)

st.title("📊 CBSE Multi-Ratio Calculator")
st.caption("Generate ratio-wise sheets with a styled Summary")

# ================= UI =================
uploaded_file = st.file_uploader("📁 Upload CBSE Excel File", type=["xlsx"])

col1, col2, col3 = st.columns(3)
with col1:
    range_input = st.text_input("🔢 Ratio Range (example 51-61)")
with col2:
    service_choice = st.selectbox("🛠 Service", ["IRIS", "FPS"])
with col3:
    final_name = st.text_input("💾 Output File Name", "cbse_multiple_ratio")

generate = st.button("🚀 Generate Excel", use_container_width=True)

# ================= LOGIC =================
if generate:

    if not uploaded_file or not range_input:
        st.error("❌ File and Ratio range are required")
        st.stop()

    try:
        start_ratio, end_ratio = map(int, range_input.split("-"))
    except:
        st.error("❌ Ratio must be like 51-61")
        st.stop()

    ratios = list(range(start_ratio, end_ratio + 1, 5))
    if end_ratio not in ratios:
        ratios.append(end_ratio)

    df = pd.read_excel(uploaded_file, sheet_name="Sheet1")

    months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    candidate_cols = [c for c in df.columns if any(m in str(c) for m in months) and "Total" not in str(c)]

    if not candidate_cols:
        st.error("❌ No candidate columns detected")
        st.stop()

    temp_df = df[candidate_cols].apply(pd.to_numeric, errors="coerce").fillna(0)
    max_candidate_sum = int(temp_df.max(axis=1).sum())

    wb = Workbook()
    wb.remove(wb.active)

    # ================= RATIO SHEETS =================
    for ratio in ratios:
        ws = wb.create_sheet(f"Ratio_{ratio}")

        for c, h in enumerate(df.columns, 1):
            ws.cell(1, c, h)

        for r, row in enumerate(df.values, 2):
            for c, v in enumerate(row, 1):
                ws.cell(r, c, v)

        col_index = {cell.value: i+1 for i, cell in enumerate(ws[1])}

        # OPR columns
        for col in candidate_cols:
            src = col_index[col]
            dst = ws.max_column + 1
            ws.cell(1, dst, f"{col} Opr")
            for r in range(2, ws.max_row + 1):
                ws.cell(r, dst, f"=ROUNDUP({ws.cell(r,src).coordinate}/{ratio},0)")

        opr_cols = [c.column for c in ws[1] if c.value and str(c.value).endswith("Opr")]

        # All-Day Max Opr
        all_day_col = ws.max_column + 1
        ws.cell(1, all_day_col, "All-Day Max Opr")
        for r in range(2, ws.max_row + 1):
            refs = ",".join(ws.cell(r,c).coordinate for c in opr_cols)
            ws.cell(r, all_day_col, f"=MAX({refs})")

        # Tab
        tab_col = ws.max_column + 1
        ws.cell(1, tab_col, "Tab")
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, all_day_col).coordinate
            ws.cell(r, tab_col,
                f"=IF({v}=0,0,IF(AND({v}>=8,{v}<16),{v}+2,IF({v}>15,{v}+3,{v}+1)))"
            )

        # Service
        service_col = ws.max_column + 1
        ws.cell(1, service_col, service_choice)
        for r in range(2, ws.max_row + 1):
            ws.cell(r, service_col, f"={ws.cell(r,tab_col).coordinate}")

        # OTG
        otg_col = ws.max_column + 1
        ws.cell(1, otg_col, "OTG")
        for r in range(2, ws.max_row + 1):
            ws.cell(r, otg_col, f"={ws.cell(r,service_col).coordinate}*2")

        total_col = col_index["Total Candidate"]

        # Hologram
        holo_col = ws.max_column + 1
        ws.cell(1, holo_col, "Hologram")
        for r in range(2, ws.max_row + 1):
            ws.cell(r, holo_col, f"=ROUNDUP({ws.cell(r,total_col).coordinate}/100,0)+1")

        # Id Card
        id_col = ws.max_column + 1
        ws.cell(1, id_col, "Id Card")
        for r in range(2, ws.max_row + 1):
            ws.cell(r, id_col, f"={ws.cell(r,all_day_col).coordinate}+1")

        # Jacket
        jacket_col = ws.max_column + 1
        ws.cell(1, jacket_col, "Jacket")
        for r in range(2, ws.max_row + 1):
            ws.cell(r, jacket_col, f"={ws.cell(r,all_day_col).coordinate}")

    # ================= SUMMARY =================
    summary = wb.create_sheet("Summary", 0)

    headers = ["Range","Total Center","Total Candidate","Max Candidate"]
    headers += [f"{c} Opr" for c in candidate_cols]
    headers += ["All-Day Max Opr","Tab",service_choice,"OTG","Hologram","Id Card","Jacket","Avg"]

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")

    for c, h in enumerate(headers, 1):
        cell = summary.cell(1, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    clean_map = {str(c).lower().strip(): i+1 for i,c in enumerate(df.columns)}
    tc_letter = get_column_letter(clean_map["total candidate"])

    base = len(df.columns) + 1

    for r, ratio in enumerate(ratios, 2):
        rs = f"Ratio_{ratio}"
        summary.cell(r,1,f"{ratio}-{ratio}")
        summary.cell(r,2,f"=COUNT('{rs}'!{tc_letter}:{tc_letter})")
        summary.cell(r,3,f"=SUM('{rs}'!{tc_letter}:{tc_letter})")
        summary.cell(r,4,max_candidate_sum)

        # OPR SUM
        for i in range(len(candidate_cols)):
            col = base + i
            summary.cell(r,5+i,f"=SUM('{rs}'!{get_column_letter(col)}:{get_column_letter(col)})")

        # Remaining SUMS
        idx = 5 + len(candidate_cols)
        for offset in range(7):
            ratio_col = base + len(candidate_cols) + offset
            summary.cell(
                r,
                idx + offset,
                f"=SUM('{rs}'!{get_column_letter(ratio_col)}:{get_column_letter(ratio_col)})"
            )

        # Avg
        summary.cell(
            r,
            len(headers),
            f"=IFERROR(ROUNDUP({summary.cell(r,4).coordinate}/{summary.cell(r,idx).coordinate},0),0)"
        )

    # ================= SAVE =================
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)

    with open(tmp.name, "rb") as f:
        st.success("✅ Excel file generated successfully")
        st.download_button(
            "⬇️ Download Excel",
            data=f,
            file_name=f"{final_name}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
